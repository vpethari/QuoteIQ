from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import Workbook

from inspect_excel import format_report, inspect_workbook, main


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def _write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> Path:
    workbook = Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = workbook.active
            ws.title = name
            first = False
        else:
            ws = workbook.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx, value)
    workbook.save(path)
    workbook.close()
    return path


def test_inspects_all_worksheets_and_column_names(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "catalog.xlsx",
        {
            "Catalog": [
                ["Part", "Description", "Qty"],
                ["A-1", "Widget", 2],
                ["B-2", "Gadget", 5],
            ],
            "Notes": [
                ["Note", "Author"],
                ["Check stock", "Pat"],
            ],
        },
    )

    result = inspect_workbook(path)

    assert result["filename"] == "catalog.xlsx"
    assert result["file_size_bytes"] > 0
    assert result["worksheet_names"] == ["Catalog", "Notes"]
    catalog = result["worksheets"][0]
    assert catalog["column_names"] == ["Part", "Description", "Qty"]
    assert catalog["used_row_count"] == 3
    assert catalog["used_column_count"] == 3
    assert catalog["data_row_count"] == 2
    assert catalog["first_10_rows"][0]["Part"] == "A-1"
    assert catalog["data_types"]["Qty"]["int"] == 2
    assert catalog["possible_header_rows"] == [1]
    notes = result["worksheets"][1]
    assert notes["column_names"] == ["Note", "Author"]


def test_string_only_data_rows_are_not_all_headers(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "all_text.xlsx",
        {
            "Sheet1": [
                ["Salsify ID", "Description"],
                ["NA1-1EEC", "120V DBL HEAD EXT CABLE"],
                ["NA1-2LB", "277V LIGHTING CABLE"],
            ]
        },
    )
    sheet = inspect_workbook(path)["worksheets"][0]
    assert sheet["possible_header_rows"] == [1]
    assert sheet["data_row_count"] == 2


def test_does_not_modify_input_file(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "readonly.xlsx",
        {"Sheet1": [["Name", "Qty"], ["Cable", 3]]},
    )
    before = _file_sha256(path)
    mtime_before = path.stat().st_mtime_ns

    inspect_workbook(path)

    assert _file_sha256(path) == before
    assert path.stat().st_mtime_ns == mtime_before


def test_reports_missing_values_blank_and_duplicate_rows(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "messy.xlsx",
        {
            "Sheet1": [
                ["Name", "Qty"],
                ["Cable", 1],
                [None, None],
                ["Cable", 1],
                ["Box", None],
            ]
        },
    )

    sheet = inspect_workbook(path)["worksheets"][0]

    assert 3 in sheet["blank_row_numbers"]
    assert sheet["blank_row_count"] == 1
    assert sheet["missing_null_by_column"]["Qty"] >= 1
    assert sheet["duplicate_row_count"] == 1
    assert sheet["duplicate_row_groups"][0]["excel_rows"] == [2, 4]


def test_detects_formulas_merged_cells_and_totals(tmp_path: Path) -> None:
    path = tmp_path / "structured.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Quote"
    ws["A1"] = "Item"
    ws["B1"] = "Amount"
    ws["A2"] = "Cable"
    ws["B2"] = 10
    ws["A3"] = "Subtotal"
    ws["B3"] = "=B2"
    ws.merge_cells("A5:B5")
    ws["A5"] = "Grand Total"
    workbook.save(path)
    workbook.close()

    sheet = inspect_workbook(path)["worksheets"][0]

    assert sheet["merged_cells"] == ["A5:B5"]
    assert sheet["formulas"]
    assert sheet["formulas"][0]["cell"] == "B3"
    assert 3 in sheet["possible_subtotal_or_total_rows"]
    assert 5 in sheet["possible_subtotal_or_total_rows"]
    assert any("merged" in item.lower() for item in sheet["unusual_formatting"])


def test_missing_file_raises(tmp_path: Path) -> None:
    missing = tmp_path / "no-such.xlsx"
    with pytest.raises(FileNotFoundError):
        inspect_workbook(missing)


def test_cli_prints_report_and_cli_missing_file(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    path = _write_workbook(
        tmp_path / "cli.xlsx",
        {"Sheet1": [["Name", "Qty"], ["Switch", 10]]},
    )

    assert main([str(path)]) == 0
    output = capsys.readouterr().out
    assert "Filename: cli.xlsx" in output
    assert "Name" in output
    assert "Qty" in output

    assert main([str(tmp_path / "missing.xlsx")]) == 2
    err = capsys.readouterr().err
    assert "not found" in err.lower()


def test_format_report_includes_required_sections(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "report.xlsx",
        {"Sheet1": [["Name", "Qty"], ["Whip", 5]]},
    )
    report = format_report(inspect_workbook(path))
    for label in (
        "Filename:",
        "File size:",
        "Worksheets",
        "Exact column names:",
        "Missing/null values by column:",
        "First 10 data rows:",
        "Last 5 data rows:",
        "Merged cells:",
        "Formulas:",
    ):
        assert label in report
