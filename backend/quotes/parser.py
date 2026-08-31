from __future__ import annotations

import warnings
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from matching.models import QuoteLine
from matching.normalizer import fold_whitespace
from matching.productcode import productcode_as_text
from matching.request_text import extract_quantity_from_text
from quotes.models import LineItem, QuoteParseError
from quotes.parse_diagnostics import record_parse_warning
from quotes.validation import (
    DESCRIPTION_ALIASES,
    PART_NUMBER_ALIASES,
    QUANTITY_ALIASES,
    is_blank,
    is_total_row,
    normalize_header,
    parse_quantity,
    resolve_optional_column,
)


def parse_quote_file(path: str | Path, source_name: str | None = None) -> list[LineItem]:
    workbook_path = Path(path)
    display_name = source_name or workbook_path.name
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Unknown extension is not supported")
            workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    except (InvalidFileException, OSError, KeyError, ValueError, zipfile.BadZipFile) as exc:
        raise QuoteParseError("Unable to read Excel workbook.") from exc

    try:
        if not workbook.sheetnames:
            raise QuoteParseError("Workbook has no worksheets.")
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        sheet_name = sheet.title
    finally:
        workbook.close()

    if not rows:
        raise QuoteParseError("Workbook is empty.")

    part_idx: int | None = None
    try:
        header_row_index, headers = _find_header_row(rows)
        description_idx = _resolve_raw_text_column(headers)
        if description_idx is None:
            raise QuoteParseError(
                "Could not identify a customer text column. Expected a header such as Name or Description."
            )
        quantity_idx = resolve_optional_column(headers, QUANTITY_ALIASES)
        part_idx = resolve_optional_column(headers, PART_NUMBER_ALIASES)
    except QuoteParseError:
        # No row looked like a labeled header -- some real quotes are just a
        # bare "description, quantity" data dump starting on row 1. Infer the
        # layout from a consistent text+number shape across sampled rows
        # instead of giving up outright.
        inferred = _infer_headerless_two_column_layout(rows)
        if inferred is None:
            raise
        description_idx, quantity_idx = inferred
        header_row_index = -1
        record_parse_warning(
            f'No header row was found in "{display_name}" -- Description and Quantity '
            "columns were inferred from the data."
        )

    items: list[LineItem] = []
    skipped_repeated_header = False
    for offset, raw in enumerate(rows[header_row_index + 1 :]):
        excel_row = header_row_index + 2 + offset
        values = list(raw or ())
        if _row_blank(values):
            continue
        if is_total_row(values):
            continue
        if quantity_idx is not None and _looks_like_repeated_header(
            _cell(values, quantity_idx), QUANTITY_ALIASES
        ):
            skipped_repeated_header = True
            continue
        raw_description = _cell(values, description_idx)
        description = "" if is_blank(raw_description) else fold_whitespace(productcode_as_text(raw_description))
        quantity = None
        if quantity_idx is not None:
            quantity = parse_quantity(_cell(values, quantity_idx), excel_row=excel_row)
        if quantity is None:
            quantity = extract_quantity_from_text(description)
        requested_part = None
        if part_idx is not None:
            raw_part = _cell(values, part_idx)
            if not is_blank(raw_part):
                requested_part = fold_whitespace(productcode_as_text(raw_part)) or None
        if not description and not requested_part:
            continue
        items.append(
            LineItem(
                source_file=display_name,
                source_sheet=sheet_name,
                source_row=excel_row,
                requested_description=description,
                quantity=quantity,
                requested_part_number=requested_part,
            )
        )
    if not items:
        raise QuoteParseError("No quote line items were found.")
    if skipped_repeated_header:
        record_parse_warning(
            f'"{display_name}" repeated a column header (e.g. "QTY") mid-file -- '
            "those rows were skipped as section labels, not line items."
        )
    return items


def line_items_to_quote_lines(items: list[LineItem]) -> list[QuoteLine]:
    return [
        QuoteLine(
            source_file=item.source_file,
            source_sheet=item.source_sheet,
            source_row=item.source_row,
            requested_description=item.requested_description,
            quantity=item.quantity,
            requested_part_number=item.requested_part_number,
        )
        for item in items
    ]


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for index, raw in enumerate(rows[:20]):
        values = list(raw or ())
        filled = [str(v).strip() for v in values if not is_blank(v)]
        if len(filled) < 1:
            continue
        headers = {
            str(value).strip(): col
            for col, value in enumerate(values)
            if not is_blank(value)
        }
        if _resolve_raw_text_column(headers) is None:
            continue
        return index, headers
    raise QuoteParseError(
        "Could not identify a customer text column in the workbook."
    )


def _resolve_raw_text_column(headers: dict[str, int]) -> int | None:
    index = resolve_optional_column(headers, DESCRIPTION_ALIASES)
    if index is not None:
        return index
    qty = resolve_optional_column(headers, QUANTITY_ALIASES)
    pn = resolve_optional_column(headers, PART_NUMBER_ALIASES)
    remaining = [col for _name, col in headers.items() if col not in {qty, pn}]
    if len(remaining) == 1:
        return remaining[0]
    if len(headers) == 1:
        return next(iter(headers.values()))
    return None


def _looks_like_repeated_header(value: object, aliases: tuple[str, ...]) -> bool:
    """True when a data-row cell is literally a column label like "QTY", not
    a real value -- some quotes repeat mini-headers before each section
    (e.g. "MV Termination" / "QTY" before a Termination sub-table)."""
    if is_blank(value) or isinstance(value, (int, float)) or isinstance(value, bool):
        return False
    return normalize_header(value) in aliases


def _infer_headerless_two_column_layout(rows: list[tuple]) -> tuple[int, int] | None:
    """Infer (description_col, quantity_col) when no row looks like a
    labeled header, but every sampled row has exactly one text value and one
    numeric value in the same two column positions -- a bare data dump with
    no header row at all. Returns None (never guesses) unless every sampled
    row agrees on both which columns are used and which one is numeric.
    """
    text_col: int | None = None
    number_col: int | None = None
    sampled = 0
    for raw in rows[:20]:
        values = list(raw or ())
        if _row_blank(values):
            continue
        filled = [(index, value) for index, value in enumerate(values) if not is_blank(value)]
        if len(filled) != 2:
            return None
        (index_a, value_a), (index_b, value_b) = filled
        a_is_number = isinstance(value_a, (int, float)) and not isinstance(value_a, bool)
        b_is_number = isinstance(value_b, (int, float)) and not isinstance(value_b, bool)
        if a_is_number == b_is_number:
            return None
        this_text_col, this_number_col = (index_b, index_a) if a_is_number else (index_a, index_b)
        if text_col is None:
            text_col, number_col = this_text_col, this_number_col
        elif (text_col, number_col) != (this_text_col, this_number_col):
            return None
        sampled += 1
    if text_col is None or sampled < 2:
        return None
    return text_col, number_col


def _cell(values: list[object], index: int) -> object:
    if index >= len(values):
        return None
    return values[index]


def _row_blank(values: list[object]) -> bool:
    return all(is_blank(value) for value in values)
