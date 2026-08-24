"""Inspect Excel workbooks without modifying them.

Usage:
    python backend/scripts/inspect_excel.py path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
import warnings
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

TOTAL_PATTERN = re.compile(
    r"\b(sub[- ]?total|grand\s+total|\btotals?\b)\b",
    re.IGNORECASE,
)


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _cell_value(cell: Cell | MergedCell) -> Any:
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def _infer_python_type(value: Any) -> str:
    if _is_empty(value):
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str):
        if str(value).startswith("="):
            return "formula"
        return "str"
    return type(value).__name__


def _row_values(ws: Worksheet, row: int, max_col: int) -> list[Any]:
    return [_cell_value(ws.cell(row, col)) for col in range(1, max_col + 1)]


def _row_is_blank(values: list[Any]) -> bool:
    return all(_is_empty(v) for v in values)


def _looks_like_header_row(values: list[Any]) -> bool:
    filled = [v for v in values if not _is_empty(v)]
    if len(filled) < 2:
        return False
    string_count = sum(1 for v in filled if isinstance(v, str) and not str(v).startswith("="))
    numeric_count = sum(1 for v in filled if isinstance(v, (int, float)) and not isinstance(v, bool))
    if numeric_count:
        return False
    if string_count != len(filled):
        return False
    # Header labels are typically short unique strings.
    if any(isinstance(v, str) and len(v) > 80 for v in filled):
        return False
    return len(set(str(v).strip().lower() for v in filled)) == len(filled)


def _looks_like_total_row(values: list[Any]) -> bool:
    for value in values:
        if isinstance(value, str) and TOTAL_PATTERN.search(value):
            return True
    return False


def inspect_worksheet(ws: Worksheet) -> dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    used_rows = 0
    for row in range(1, max_row + 1):
        if not _row_is_blank(_row_values(ws, row, max_col)):
            used_rows = row

    effective_rows = used_rows
    effective_cols = 0
    for col in range(1, max_col + 1):
        for row in range(1, effective_rows + 1):
            if not _is_empty(_cell_value(ws.cell(row, col))):
                effective_cols = col
                break

    max_row = effective_rows
    max_col = effective_cols

    merged = [str(rng) for rng in ws.merged_cells.ranges]
    formulas: list[dict[str, Any]] = []
    unusual: list[str] = []

    if merged:
        unusual.append(f"{len(merged)} merged cell range(s)")
    if ws.freeze_panes:
        unusual.append(f"frozen panes at {ws.freeze_panes}")
    if ws.auto_filter and ws.auto_filter.ref:
        unusual.append(f"auto-filter {ws.auto_filter.ref}")
    if ws.sheet_state != "visible":
        unusual.append(f"sheet_state={ws.sheet_state}")

    all_rows: list[list[Any]] = []
    blank_rows: list[int] = []
    possible_headers: list[int] = []
    possible_totals: list[int] = []
    header_signature: list[Any] | None = None

    for row in range(1, max_row + 1):
        values = _row_values(ws, row, max_col)
        all_rows.append(values)
        if _row_is_blank(values):
            blank_rows.append(row)
            continue
        signature = [_normalize_for_dup(v) for v in values]
        if header_signature is None and _looks_like_header_row(values):
            possible_headers.append(row)
            header_signature = signature
        elif header_signature is not None and signature == header_signature:
            possible_headers.append(row)
        if _looks_like_total_row(values):
            possible_totals.append(row)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                formulas.append(
                    {
                        "cell": cell.coordinate,
                        "formula": value,
                    }
                )
            if cell.number_format not in ("General", "0", "@") and not _is_empty(value):
                unusual.append(
                    f"{cell.coordinate} number_format={cell.number_format!r}"
                )

    header_row = possible_headers[0] if possible_headers else (1 if max_row else None)
    if header_row:
        column_names = [
            str(v).strip() if not _is_empty(v) else f"Column{get_column_letter(i)}"
            for i, v in enumerate(_row_values(ws, header_row, max_col), start=1)
        ]
        data_start = header_row + 1
    else:
        column_names = [f"Column{get_column_letter(i)}" for i in range(1, max_col + 1)]
        data_start = 1

    missing = {name: 0 for name in column_names}
    type_counts = {name: Counter() for name in column_names}
    data_rows: list[dict[str, Any]] = []
    value_tuples: list[tuple[Any, ...]] = []

    for excel_row, values in enumerate(all_rows, start=1):
        if excel_row < data_start:
            continue
        if excel_row in blank_rows:
            for name in column_names:
                missing[name] += 1
            continue
        record = {"_excel_row": excel_row}
        normalized: list[Any] = []
        for name, value in zip(column_names, values):
            record[name] = value
            normalized.append(value)
            if _is_empty(value):
                missing[name] += 1
                type_counts[name]["empty"] += 1
            else:
                type_counts[name][_infer_python_type(value)] += 1
        data_rows.append(record)
        value_tuples.append(tuple(_normalize_for_dup(v) for v in normalized))

    dup_counts = Counter(value_tuples)
    duplicate_groups = []
    seen: set[tuple[Any, ...]] = set()
    for record, key in zip(data_rows, value_tuples):
        if dup_counts[key] > 1 and key not in seen:
            seen.add(key)
            rows_for_key = [
                r["_excel_row"] for r, k in zip(data_rows, value_tuples) if k == key
            ]
            duplicate_groups.append(
                {"count": dup_counts[key], "excel_rows": rows_for_key, "values": list(key)}
            )

    first10 = data_rows[:10]
    last5 = data_rows[-5:] if data_rows else []

    data_types = {
        name: dict(counter) if counter else {"empty": 0}
        for name, counter in type_counts.items()
    }

    # Deduplicate noisy number-format notes while keeping order.
    unusual_unique: list[str] = []
    for item in unusual:
        if item not in unusual_unique:
            unusual_unique.append(item)

    return {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "used_row_count": max_row,
        "used_column_count": max_col,
        "column_names": column_names,
        "header_row": header_row,
        "possible_header_rows": possible_headers,
        "possible_subtotal_or_total_rows": possible_totals,
        "blank_row_numbers": blank_rows,
        "blank_row_count": len(blank_rows),
        "data_row_count": len(data_rows),
        "duplicate_row_groups": duplicate_groups,
        "duplicate_row_count": sum(g["count"] - 1 for g in duplicate_groups),
        "missing_null_by_column": missing,
        "data_types": data_types,
        "first_10_rows": first10,
        "last_5_rows": last5,
        "merged_cells": merged,
        "formulas": formulas,
        "unusual_formatting": unusual_unique,
        "freeze_panes": ws.freeze_panes,
        "auto_filter": ws.auto_filter.ref if ws.auto_filter else None,
        "sheet_state": ws.sheet_state,
        "dimensions": ws.dimensions,
    }


def _normalize_for_dup(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def inspect_workbook(path: str | Path) -> dict[str, Any]:
    excel_path = Path(path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not excel_path.is_file():
        raise FileNotFoundError(f"Not a file: {excel_path}")

    load_warnings: list[str] = []
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook: Workbook = load_workbook(excel_path, data_only=False, read_only=False)
        for warning in caught:
            load_warnings.append(str(warning.message))

    try:
        sheets = [inspect_worksheet(ws) for ws in workbook.worksheets]
    finally:
        workbook.close()

    return {
        "filename": excel_path.name,
        "path": str(excel_path.resolve()),
        "file_size_bytes": excel_path.stat().st_size,
        "worksheet_names": [s["name"] for s in sheets],
        "worksheet_count": len(sheets),
        "load_warnings": load_warnings,
        "worksheets": sheets,
    }


def format_report(result: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"Filename: {result['filename']}")
    lines.append(f"Path: {result['path']}")
    lines.append(f"File size: {result['file_size_bytes']} bytes")
    lines.append(f"Worksheets ({result['worksheet_count']}): {', '.join(result['worksheet_names'])}")
    if result["load_warnings"]:
        lines.append("Load warnings:")
        for warning in result["load_warnings"]:
            lines.append(f"  - {warning}")
    else:
        lines.append("Load warnings: none")

    for sheet in result["worksheets"]:
        lines.append("")
        lines.append(f"=== Worksheet: {sheet['name']} ===")
        lines.append(f"Declared dimensions: {sheet['dimensions']}")
        lines.append(f"Rows (used): {sheet['used_row_count']}")
        lines.append(f"Columns (used): {sheet['used_column_count']}")
        lines.append(f"openpyxl max_row/max_column: {sheet['max_row']}/{sheet['max_column']}")
        lines.append(f"Exact column names: {sheet['column_names']}")
        lines.append(f"Possible header rows: {sheet['possible_header_rows'] or 'none'}")
        lines.append(f"Chosen header row: {sheet['header_row']}")
        lines.append(f"Data rows: {sheet['data_row_count']}")
        lines.append(f"Blank rows: {sheet['blank_row_count']} {sheet['blank_row_numbers']}")
        lines.append(
            "Possible subtotal/total rows: "
            f"{sheet['possible_subtotal_or_total_rows'] or 'none'}"
        )
        lines.append("Data types by column (value counts):")
        for name, counts in sheet["data_types"].items():
            lines.append(f"  - {name}: {dict(counts)}")
        lines.append("Missing/null values by column:")
        for name, count in sheet["missing_null_by_column"].items():
            lines.append(f"  - {name}: {count}")
        lines.append(f"Duplicate extra rows: {sheet['duplicate_row_count']}")
        if sheet["duplicate_row_groups"]:
            for group in sheet["duplicate_row_groups"]:
                lines.append(
                    f"  - rows {group['excel_rows']} repeated {group['count']} times"
                )
        lines.append("First 10 data rows:")
        if not sheet["first_10_rows"]:
            lines.append("  (none)")
        for row in sheet["first_10_rows"]:
            lines.append(f"  {row}")
        lines.append("Last 5 data rows:")
        if not sheet["last_5_rows"]:
            lines.append("  (none)")
        for row in sheet["last_5_rows"]:
            lines.append(f"  {row}")
        lines.append(f"Merged cells: {sheet['merged_cells'] or 'none'}")
        lines.append(f"Formulas: {sheet['formulas'] or 'none'}")
        lines.append(
            "Unusual formatting / structure: "
            f"{sheet['unusual_formatting'] or 'none'}"
        )
        lines.append(f"Freeze panes: {sheet['freeze_panes']}")
        lines.append(f"Auto filter: {sheet['auto_filter']}")
        lines.append(f"Sheet state: {sheet['sheet_state']}")

    return "\n".join(lines) + "\n"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspect an Excel workbook without modifying it."
    )
    parser.add_argument(
        "excel_path",
        help="Path to an .xlsx/.xlsm workbook",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        result = inspect_workbook(args.excel_path)
    except FileNotFoundError as exc:
        print(str(exc), file=sys.stderr)
        return 2
    except Exception as exc:  # pragma: no cover - unexpected openpyxl errors
        print(f"Failed to inspect workbook: {exc}", file=sys.stderr)
        return 1
    print(format_report(result), end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
