from __future__ import annotations

import warnings
from pathlib import Path

from openpyxl import load_workbook

from catalog.classify import classify_row
from catalog.normalize import as_text, normalize_description, normalize_whitespace
from catalog.parsing import CatalogNumberParseError, parse_catalog_number
from matching.models import ProductRecord, QuoteLine

COL_SALSIFY = "Salsify ID"
COL_CATALOG = "Catalog Number - Short Description"
COL_DESCRIPTION = "Short Description - en-US"
COL_PARENT = "salsify:parent_id"


def _header_map(values: list[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        if value is None:
            continue
        mapping[str(value).strip()] = index
    return mapping


def load_catalog_records(path: str | Path) -> list[ProductRecord]:
    """Load product and family records from the Atkore Excel extract."""
    workbook_path = Path(path)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Unknown extension is not supported")
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    headers = _header_map(list(rows[0]))
    required = (COL_SALSIFY, COL_CATALOG, COL_DESCRIPTION, COL_PARENT)
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"{workbook_path.name} is missing columns: {missing}")

    records: list[ProductRecord] = []
    for raw in rows[1:]:
        values = list(raw)
        salsify = as_text(values[headers[COL_SALSIFY]])
        combined = as_text(values[headers[COL_CATALOG]])
        description = as_text(values[headers[COL_DESCRIPTION]])
        parent = as_text(values[headers[COL_PARENT]])
        kind = classify_row(salsify, combined, description, parent)
        if kind == "invalid" or not salsify:
            continue
        official = None
        if kind == "product":
            try:
                official = parse_catalog_number(combined)
            except CatalogNumberParseError:
                continue
        records.append(
            ProductRecord(
                salsify_id=salsify.strip(),
                official_part_number=official,
                description=normalize_description(description),
                record_type=kind,
                parent_id=normalize_whitespace(parent),
                catalog_number_and_description=normalize_whitespace(combined),
            )
        )
    return records


def load_matchable_products(path: str | Path) -> list[ProductRecord]:
    return [item for item in load_catalog_records(path) if item.record_type == "product"]


from quotes.parser import line_items_to_quote_lines, parse_quote_file


def load_quote_lines(path: str | Path) -> list[QuoteLine]:
    return line_items_to_quote_lines(parse_quote_file(path))
